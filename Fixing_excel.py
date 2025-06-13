import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime

# --- CONFIGURATION ---
target_date = '12-06-2025'
DOWNLOAD_DIR = r'C:\Users\91987\TVS\downloads'
combined_file = os.path.join(DOWNLOAD_DIR, f'Combined_Report_{target_date}.xlsx')
output_file = os.path.join(DOWNLOAD_DIR, f'Processed_Combined_Report_{target_date}.xlsx')

# Overwrite output file if it exists
if os.path.exists(output_file):
    os.remove(output_file)

columns_to_keep = [
    'Ticket Number', 'Call Log Date', 'Call Log Time',
    'Actual Response/Reach Date as per Dealer', 'Actual Response/Reach Time as per Dealer',
    'Response/Reach Gap', 'Actual Restoration Date Dealer', 'Actual Restoration Time Dealer',
    'Total Restoration Time', 'Company Name', 'Registration Number', 'Chassis Number',
    'Customer Type', 'Restoration Type', 'Estimated Response/Reach Time'
]

# List of national holidays (add more as needed)
national_holidays = [
    '2025-01-26',  # Republic Day
    '2025-03-14',  # Holi (2025)
    '2025-10-20',  # Diwali Day 1 (2025)
    '2025-10-21',  # Diwali Day 2 (2025)
    '2025-10-22',  # Diwali Day 3 (2025)
    '2025-08-15',  # Independence Day
    '2025-10-02',  # Gandhi Jayanti
]

def is_holiday(date_obj):
    return date_obj.weekday() == 6 or date_obj.strftime('%Y-%m-%d') in national_holidays

def get_quarter(month):
    if month in [4, 5, 6]:
        return 'Q1'
    elif month in [7, 8, 9]:
        return 'Q2'
    elif month in [10, 11, 12]:
        return 'Q3'
    else:
        return 'Q4'

def auto_fit_excel(filename):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(filename)

with pd.ExcelFile(combined_file) as xls, pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        df = df[[col for col in columns_to_keep if col in df.columns]]
        df = df[df['Restoration Type'] == 'Restored By Support'].copy()
        if df.empty:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            continue
        # Month column in 'Month Year' format, but keep date parsing in dd-mm-yyyy
        df['Month'] = pd.to_datetime(df['Call Log Date'], format='%d-%m-%Y', errors='coerce').dt.strftime('%B %y')
        # Date Time (TTBL) in dd-mm-yyyy HH:MM:SS
        df['Date Time (TTBL)'] = pd.to_datetime(
            df['Call Log Date'].astype(str) + ' ' + df['Call Log Time'].astype(str), format='%d-%m-%Y %H:%M:%S', errors='coerce')
        df['Date Time (TTBL)'] = df['Date Time (TTBL)'].dt.strftime('%d-%m-%Y %H:%M:%S')
        # Date Time (Dealer) in dd-mm-yyyy HH:MM:SS
        df['Date Time (Dealer)'] = pd.to_datetime(
            df['Actual Response/Reach Date as per Dealer'].astype(str) + ' ' +
            df['Actual Response/Reach Time as per Dealer'].astype(str), format='%d-%m-%Y %H:%M:%S', errors='coerce')
        df['Date Time (Dealer)'] = df['Date Time (Dealer)'].dt.strftime('%d-%m-%Y %H:%M:%S')
        # Restored as per Dealer in dd-mm-yyyy HH:MM:SS
        df['Restored as per Dealer'] = df['Actual Restoration Date Dealer'].astype(str) + ' ' + df['Actual Restoration Time Dealer'].astype(str)
        df['Restored as per Dealer'] = pd.to_datetime(df['Restored as per Dealer'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
        df['Restored as per Dealer'] = df['Restored as per Dealer'].dt.strftime('%d-%m-%Y %H:%M:%S')
        # For calculations, parse datetimes again
        dt_ttbl = pd.to_datetime(df['Date Time (TTBL)'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
        dt_dealer = pd.to_datetime(df['Date Time (Dealer)'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
        dt_restored = pd.to_datetime(df['Restored as per Dealer'], format='%d-%m-%Y %H:%M:%S', errors='coerce')
        # Response Time (in hours:minutes)
        resp_seconds = (dt_dealer - dt_ttbl).dt.total_seconds()
        df['Response Time'] = resp_seconds.apply(lambda x: f"{int(x//3600):02}:{int((x%3600)//60):02}" if pd.notnull(x) else '')
        # Restoration Time (in hours:minutes)
        rest_seconds = (dt_restored - dt_ttbl).dt.total_seconds()
        df['Restoration Time'] = rest_seconds.apply(lambda x: f"{int(x//3600):02}:{int((x%3600)//60):02}" if pd.notnull(x) else '')
        # For confirmity checks, use hours as float
        resp_hours = resp_seconds / 3600
        rest_hours = rest_seconds / 3600
        df['Response Confirmity (2 Hrs)'] = resp_hours.apply(lambda x: 'Conf.' if pd.notnull(x) and x <= 2 else 'NC')
        df['Response Confirmity (4 Hrs)'] = resp_hours.apply(lambda x: 'Conf.' if pd.notnull(x) and x <= 4 else 'NC')
        df['Restore Confirmity'] = rest_hours.apply(lambda x: 'Conf.' if pd.notnull(x) and x <= 12 else 'NC')
        df['Holiday Count'] = dt_ttbl.apply(lambda x: is_holiday(x) if pd.notnull(x) else False)
        df['Day/Night'] = dt_ttbl.apply(
            lambda x: 'Day' if pd.notnull(x) and 6 <= x.hour < 18 else ('Night' if pd.notnull(x) else ''))
        df['Quarter'] = dt_ttbl.apply(
            lambda x: get_quarter(x.month) if pd.notnull(x) else '')
        df.to_excel(writer, sheet_name=sheet_name, index=False)
auto_fit_excel(output_file)
print(f"Processed file saved as: {output_file}") 